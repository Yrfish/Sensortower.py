import time
from random import randrange
from auth_data import password, email
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import Font

wb = load_workbook(filename="Data.xlsx", data_only=True)
ws = wb.active
urls = list()

for row in range(3, 1222):
    if ws["G" + f"{row}"].value == None:
        ios = ws["D" + f"{row}"].value
        andr = ws["E" + f"{row}"].value
        urs = (ios, andr)
        urls.append(urs)

    index_list = list()

    for i in range(3, 1222):
        if ws["G" + f"{i}"].value == None:
            index_list.append(i)

print(index_list)
print(f"Get {len(urls)} url from excel")
print(urls)

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-automation'])
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)


def login_in(email, password):
    # login in
    print("\n Now i am start parsing")
    driver.get("https://app.sensortower.com/users/sign_in?return_to_path=%2Fusers%2Fsign_up")

    time.sleep(3)
    email_inp = driver.find_element(By.NAME, "user[email]")
    email_inp.clear()
    email_inp.send_keys(email)
    time.sleep(3)
    password_inp = driver.find_element(By.NAME, "user[password]")
    password_inp.clear()
    password_inp.send_keys(password)
    time.sleep(3)
    password_inp.send_keys(Keys.ENTER)
    time.sleep(5)


def parse(urls):
    login_in(email, password)
    iteration = 0
    index = 0  # get index in index list
    font = Font(bold=True)

    for ios, andr in urls:
        time.sleep(3)
        try:
            try:
                driver.get(ios)
                time.sleep(randrange(4, 8))
                app_link = driver.find_element(By.XPATH,
                                               "/html/body/div[3]/div/div[3]/table/tbody/tr[1]/td[3]/div/div/a[1]")
                link = app_link.get_property("href")
                print(link)

            except:
                driver.get(andr)
                time.sleep(randrange(4, 8))
                app_link = driver.find_element(By.XPATH,
                                               "/html/body/div[3]/div/div[3]/table/tbody/tr[1]/td[3]/div/div/a[1]")
                link = app_link.get_property("href")
                print(link)
        except:
            ws[f"G{index_list[index]}"] = "does not exist"
            ws[f"G{index_list[index]}"].font = font
            print("does not exist")
            index += 1
            iteration += 1
            continue

        try:
            app_id = link.split("/")[8]
            print("app id is - ", app_id)
        except:
            app_id = link.split("/")[4]
            print("app id is - ", app_id)

        driver.get(link)
        time.sleep(10)
        try:
            support_url = driver.find_element(By.XPATH,
                                              "/html/body/div[3]/div/div/div/div/div[1]/section/div/div[2]/div[2]/div[1]/div/div/a").get_attribute(
                "href")
        except:
            support_url = "-"

        try:
            subtitle = driver.find_element(By.XPATH,
                                           "/html/body/div[3]/div/div/div/div/div[1]/div[2]/div/div[2]/div[1]/div[1]/div/div/div/div/div/div[2]/div[1]/div").text
        except:
            subtitle = "does not exist for this app"

        divs_country = driver.find_element(By.XPATH,
                                           "/html/body/div[3]/div/div/div/div/div[1]/section/div/div[2]/div[1]/div[3]/div/div")
        span_countries = divs_country.find_elements(By.TAG_NAME, "span")
        countries = list()

        for span in span_countries:
            country = span.text
            countries.append(country)
            top_countries = "".join(countries)
            print(top_countries)
        if span_countries == []:
            top_countries = "N/A"
        ps = driver.find_elements(By.TAG_NAME, "p")
        for p in ps:
            if "English" in p.text:
                languages = p.text
            elif not "English" in p.text:
                languages = "N/A"

        try:

            divs = driver.find_elements(By.CLASS_NAME, "css-19cssbn")

            if len(divs) == 2:
                div = divs[1]
                description = div.find_element(By.TAG_NAME, "div").text
            else:
                div = divs[0]
                description = div.find_element(By.TAG_NAME, "div").text
        except:
            description = "N/A"

        dev_link = driver.find_element(By.XPATH,
                                       "/html/body/div[3]/div/div/div/div/div[1]/section/div/div[1]/div[1]/div/div/a[1]").get_attribute(
            "href")
        dev_id = dev_link.split("/")[-1]
        driver.get(dev_link)
        time.sleep(6)
        download_dev = driver.find_element(By.XPATH, "/html/body/div[3]/div[3]/div[2]/h3/a").text
        revenue_dev = driver.find_element(By.XPATH, "/html/body/div[3]/div[3]/div[3]/h3/a").text
        publisher_summary = driver.find_element(By.XPATH, "/html/body/div[3]/div[4]/div").text


        for j in range(0, 7):
            print("я скролю")
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(randrange(2, 3))

        tr_tags = driver.find_elements(By.TAG_NAME, "tr")
        for tr_tag in tr_tags:
            if tr_tag.get_attribute("data-entity-id") == app_id:  # find exact row for app in table
                print(tr_tag.get_attribute("data-entity-id"))
                app_row = tr_tag

        time.sleep(3)
        tags_a = app_row.find_elements(By.TAG_NAME, "a")  # in app row find a tag, a tag contains needed information

        for tag_a in tags_a:
            if tag_a.get_attribute(
                    "data-bind") == "text: $data.worldwideLastMonthDownloads.string, click: function() { $parent.activateDemoFormModal('App Downloads') }":  # Check if it is extact tag that contains amount of download
                app_down = tag_a.text

            elif tag_a.get_attribute(
                    "data-bind") == "text: $data.worldwideLastMonthRevenue.string, click: function() { $parent.activateDemoFormModal('App Revenue') }":  # Check if it is extact tag that contains app revenue
                app_rev = tag_a.text
                print("revenue", app_rev)

        print(dev_link)
# hardly writing information on each iteration and save it (saving get more tome but prevent missing data if we have an error)
        ws[f"G{index_list[index]}"] = app_id
        ws[f"H{index_list[index]}"] = dev_id
        ws[f"K{index_list[index]}"] = description
        ws[f"L{index_list[index]}"] = subtitle
        ws[f"M{index_list[index]}"] = languages
        ws[f"O{index_list[index]}"] = download_dev
        ws[f"P{index_list[index]}"] = revenue_dev
        ws[f"R{index_list[index]}"] = support_url
        ws[f"S{index_list[index]}"] = publisher_summary
        ws[f"J{index_list[index]}"] = app_rev
        ws[f"I{index_list[index]}"] = app_down
        ws[f"Q{index_list[index]}"] = top_countries
        print(f"stop on row - {index_list[index]}")
        wb.save("Data1.xlsx")
        print(f"iteration - {iteration}")

        index += 1

        if iteration == 10:  # sleep each 10 iteration to reset requests timer
            iteration = 0
            print("snooooozee 60sec >>>> zzz....zz..zzzz.zz")
            time.sleep(60)
            print("\n i have waked up, saving file)))")
            wb.save("Data1.xlsx")

        iteration += 1

        time.sleep(2)

    wb.save("Data1.xlsx")
    driver.quit()
    driver.close()
    print("finish data collection")


parse(urls)
print("finish data collection")
print("finish data collection")
print("file 'Data1.xlsx' save in current directrory")
